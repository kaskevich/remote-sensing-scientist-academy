#!/usr/bin/env python3
"""Build the Species Atlas study layer and refresh official FinBIF metadata.

The source workbooks are read-only inputs and are never copied into the public
site. FINBIF_ACCESS_TOKEN is used with api.laji.fi when available; otherwise the
same public laji.fi proxy used by the human-readable taxon pages is queried.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import statistics
import time
import urllib.parse
import urllib.request
from io import BytesIO
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PIL import Image


ROOT = Path(__file__).resolve().parents[2]
CONTENT = ROOT / "content" / "species"
DATA = ROOT / "data" / "species"
DOCS = ROOT / "docs" / "species"
HABITATS = ("OP", "LS", "US", "TG")
NON_TAXA = {"Bare_ground", "Litter", "Moss"}
SITES = ("Saardu", "Keemu", "Koera", "Kudani")
RETRIEVED_AT = date.today().isoformat()

# These decisions are explicit because fuzzy string results are never accepted.
# The target itself must resolve as an exact current FinBIF scientific name.
MANUAL_TARGETS = {
    "Agrostis_capilaris": ("Agrostis capillaris", "spelling-normalization", "One-letter spelling correction."),
    "Arrhenaterum_elatius": ("Arrhenatherum elatius", "spelling-normalization", "Missing h in the study label."),
    "Carex_panaceae": ("Carex panicea", "manual-review", "Probable spelling variant supported by the supplied seed taxon, but the label is ambiguous and remains flagged."),
    "Festuca_arundinacea": ("Lolium arundinaceum", "synonym-match", "FinBIF accepted-name search links the historical Festuca name to Lolium."),
    "Galium_borealium": ("Galium boreale", "spelling-normalization", "Study epithet normalized to the verified FinBIF name."),
    "Glaux_maritima": ("Lysimachia maritima", "synonym-match", "FinBIF returns Glaux maritima as a synonym."),
    "Inula_salicina": ("Pentanema salicinum", "synonym-match", "FinBIF returns Inula salicina as a synonym."),
    "Juncus_gerardii": ("Juncus gerardi", "spelling-normalization", "Study epithet has an extra i."),
    "Leontodon_autumnalis": ("Scorzoneroides autumnalis", "synonym-match", "FinBIF returns the historical Leontodon name as a synonym."),
    "Potentilla_anserina": ("Argentina anserina", "synonym-match", "FinBIF returns the historical Potentilla name as a synonym."),
    "Potentilla_erectus": ("Potentilla erecta", "spelling-normalization", "Study epithet corrected to the verified feminine form."),
    "Salicornia_europaea": ("Salicornia perennans", "manual-review", "The study aggregate name is linked to the supplied seed taxon for display, but the Salicornia complex requires expert review."),
    "Sesleria_caerulea": ("Sesleria uliginosa", "synonym-match", "FinBIF accepted-name search links the study name to Sesleria uliginosa."),
    "Stelaria_graminea": ("Stellaria graminea", "spelling-normalization", "Missing l in the study genus."),
}

LICENSES = {
    "CC-BY-4.0": ("CC BY 4.0", "https://creativecommons.org/licenses/by/4.0/"),
    "CC-BY-SA-4.0": ("CC BY-SA 4.0", "https://creativecommons.org/licenses/by-sa/4.0/"),
    "CC-BY-NC-4.0": ("CC BY-NC 4.0", "https://creativecommons.org/licenses/by-nc/4.0/"),
    "CC-BY-NC-SA-4.0": ("CC BY-NC-SA 4.0", "https://creativecommons.org/licenses/by-nc-sa/4.0/"),
    "CC0-1.0": ("CC0 1.0", "https://creativecommons.org/publicdomain/zero/1.0/"),
}

# Primary card images were selected after a visual audit of all 78 current
# primaries. These IDs favour living growth form, reproductive morphology or a
# clear diagnostic structure; licensing is still revalidated on every import.
PRIMARY_MEDIA_IDS = {
    "MX.38844": "MM.121803",   # Argentina anserina, flowering living plant
    "MX.40557": "MM.80310",    # Arrhenatherum elatius, flowering stems
    "MX.40564": "MM.112202",   # Deschampsia cespitosa, living tussock
    "MX.40537": "MM.124143",   # Elytrigia repens, flowering stand
    "MX.39126": "MM.125633",   # Linum catharticum, flowering plant
    "MX.39057": "MM.3872258",  # Lotus corniculatus, flowers and foliage
    "MX.40639": "MM.122065",   # Phragmites australis, living reed heads
    "MX.39038": "MM.123004",   # Trifolium repens, flowering living stand
    "MX.38073": "MM.110495",   # Cerastium fontanum, living foliage
    "MX.37718": "MM.388270",   # Equisetum palustre, diagnostic strobili
    "MX.39959": "MM.79260",    # Pilosella officinarum, flowering plant
    "MX.39199": "MM.77611",    # Pimpinella saxifraga, flowering stand
    "MX.39500": "MM.106710",   # Prunella vulgaris, flowering inflorescence
    "MX.38279": "MM.122256",   # Rumex acetosa, living rosette
    "MX.39042": "MM.177296",   # Trifolium fragiferum, flower and fruit
    "MX.39052": "MM.122991",   # Trifolium pratense, flowering living stand
    "MX.40457": "MM.108931",   # Poa annua, diagnostic inflorescence
    "MX.39673": "MM.2208050",  # Plantago major, flowering rosette
    "MX.39917": "MM.125598",   # Scorzoneroides autumnalis, flowering plant
    "MX.38972": "MM.112197",   # Vicia cracca, living growth form
    "MX.39366": "MM.3874530",  # Succisa pratensis, diagnostic flower
    "MX.40643": "MM.105744",   # Molinia caerulea, living stand
    "MX.39284": "MM.125483",   # Galium boreale, flowers and foliage
    "MX.39002": "MM.125579",   # Lathyrus pratensis, diagnostic flower
    "MX.40461": "MM.122051",   # Poa pratensis, flowering stand
}


def fetch_json(path: str, query: dict[str, str] | None = None) -> Any:
    token = os.getenv("FINBIF_ACCESS_TOKEN")
    params = dict(query or {})
    if token:
        params["access_token"] = token
        base = "https://api.laji.fi/v0"
    else:
        base = "https://laji.fi/api"
    url = f"{base}/{path.lstrip('/')}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    request = urllib.request.Request(url, headers={"User-Agent": "Remote-Sensing-Scientist-Academy/1.0"})
    last_error: Exception | None = None
    for attempt in range(4):
        try:
            with urllib.request.urlopen(request, timeout=45) as response:
                return json.load(response)
        except Exception as error:  # pragma: no cover - network maintenance path
            last_error = error
            time.sleep(2**attempt)
    raise RuntimeError(f"FinBIF request failed: {url}") from last_error


def exact_match(query: str) -> dict[str, Any] | None:
    results = fetch_json("taxa/search", {"query": query})
    matches = [item for item in results if item.get("type") == "exactMatches"]
    return matches[0] if matches else None


def resolve_name(study_name: str) -> dict[str, Any]:
    normalized = study_name.replace("_", " ")
    direct = exact_match(normalized)
    if direct:
        status = "synonym-match" if direct.get("nameType") == "MX.hasSynonym" else "exact"
        return reconciliation_record(study_name, normalized, direct, status, "finbif-exact-search", None)
    if study_name in MANUAL_TARGETS:
        target, status, note = MANUAL_TARGETS[study_name]
        verified = exact_match(target)
        if verified and verified.get("scientificName") == target:
            return reconciliation_record(study_name, normalized, verified, status, "reviewed-target-with-finbif-exact-search", note)
    return {
        "studyName": study_name,
        "normalizedStudyName": normalized,
        "finbifAcceptedName": None,
        "finbifScientificNameAuthorship": None,
        "taxonId": None,
        "taxonRank": None,
        "taxonomicStatus": None,
        "synonyms": [],
        "matchStatus": "unresolved" if normalized.lower().endswith(" sp") else "manual-review",
        "matchMethod": "no-exact-finbif-match",
        "matchNotes": "No exact FinBIF name or synonym match was accepted; fuzzy results were rejected.",
    }


def reconciliation_record(study_name: str, normalized: str, match: dict[str, Any], status: str, method: str, note: str | None) -> dict[str, Any]:
    matching_name = match.get("matchingName")
    accepted = match.get("scientificName")
    synonyms = [matching_name] if matching_name and accepted and matching_name != accepted else []
    return {
        "studyName": study_name,
        "normalizedStudyName": normalized,
        "finbifAcceptedName": accepted,
        "finbifScientificNameAuthorship": match.get("scientificNameAuthorship"),
        "taxonId": match.get("id"),
        "taxonRank": (match.get("taxonRank") or "").removeprefix("MX.") or None,
        "taxonomicStatus": "accepted" if match.get("nameType") == "MX.scientificName" else "accepted name with matched synonym",
        "synonyms": synonyms,
        "commonNames": match.get("vernacularName") or {},
        "matchStatus": status,
        "matchMethod": method,
        "matchNotes": note,
    }


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def quartiles(values: list[float]) -> tuple[float, float]:
    if len(values) == 1:
        return values[0], values[0]
    cuts = statistics.quantiles(values, n=4, method="inclusive")
    return cuts[0], cuts[2]


def summarize_values(values: list[float]) -> dict[str, float | int | None]:
    if not values:
        return {"median": None, "q1": None, "q3": None, "mean": None, "max": None, "n": 0}
    q1, q3 = quartiles(values)
    return {
        "median": round(statistics.median(values), 3),
        "q1": round(q1, 3),
        "q3": round(q3, 3),
        "mean": round(statistics.mean(values), 3),
        "max": round(max(values), 3),
        "n": len(values),
    }


def load_study(workbook_path: Path) -> tuple[list[str], dict[str, Any], list[dict[str, Any]], dict[str, Any]]:
    book = load_workbook(workbook_path, read_only=True, data_only=True)
    total_rows = list(book["inwork_TOTAL_SPECIES"].iter_rows(min_row=2, values_only=True))
    study_names = [row[0] for row in total_rows if row[0] and row[0] not in NON_TAXA]

    sheet = book["Community_level_data"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    records = [dict(zip(headers, row)) for row in rows[1:]]
    cover_names = [name for name in headers[19:] if name not in NON_TAXA]
    plot_totals = Counter(record["plantcommunity"] for record in records)
    site_totals = Counter(record["site"] for record in records)

    trait_tables: dict[str, dict[str, Any]] = {}
    for key, sheet_name, prefix, unit in (
        ("CCI", "CCI_medians_perSpecies", "CCI", "index"),
        ("LA", "LA_medians_perSpecies", "LA", "cm²"),
    ):
        trait_sheet = book[sheet_name]
        trait_rows = list(trait_sheet.iter_rows(values_only=True))
        trait_headers = list(trait_rows[0])
        trait_tables[key] = {
            row[0]: {
                "median": row[1], "min": row[2], "max": row[3], "average": row[4], "n": row[5],
                "unit": unit, "sourceSheet": sheet_name,
                "sourceMinHeader": trait_headers[2],
            }
            for row in trait_rows[1:] if row[0] and number(row[5]) is not None and number(row[5]) >= 5
        }

    summaries: dict[str, Any] = {}
    for name in study_names:
        habitats: dict[str, Any] = {}
        sites: dict[str, Any] = {}
        if name in cover_names:
            for code in HABITATS:
                subset = [record for record in records if record["plantcommunity"] == code]
                occupied_values = [value for record in subset if (value := number(record.get(name))) is not None and value > 0]
                occupied = len(occupied_values)
                total = plot_totals[code]
                habitats[code] = {
                    "sampled": True,
                    "occupiedPlots": occupied,
                    "totalPlots": total,
                    "occurrenceFrequency": round(occupied / total, 6),
                    "coverAmongOccupiedPlots": summarize_values(occupied_values),
                    "zeroPlotsExcludedFromCoverSummary": True,
                }
            for site in SITES:
                subset = [record for record in records if record["site"] == site]
                occupied = sum(1 for record in subset if (number(record.get(name)) or 0) > 0)
                sites[site] = {
                    "recorded": occupied > 0,
                    "occupiedPlots": occupied,
                    "totalPlots": site_totals[site],
                    "occurrenceFrequency": round(occupied / site_totals[site], 6),
                }
        summaries[name] = {
            "studyName": name,
            "coverAvailable": name in cover_names,
            "habitats": habitats,
            "sites": sites,
            "occupiedPlots": sum(item["occupiedPlots"] for item in habitats.values()),
            "totalPlots": len(records) if name in cover_names else None,
            "traits": {key: table[name] for key, table in trait_tables.items() if name in table},
        }

    audit: list[dict[str, Any]] = []
    sample_counts = Counter(record["SampleID"] for record in records)
    for sample_id, count in sample_counts.items():
        if count > 1:
            audit.append(issue("duplicate SampleID", "Community_level_data", None, sample_id, "error", f"{count} rows share this ID", False))
    for record in records:
        if record["plantcommunity"] not in HABITATS:
            audit.append(issue("invalid habitat code", "Community_level_data", None, record["SampleID"], "error", str(record["plantcommunity"]), False))
        for name in cover_names:
            value = record.get(name)
            parsed = number(value)
            if value not in (None, "") and parsed is None:
                audit.append(issue("nonnumeric cover", "Community_level_data", name, record["SampleID"], "error", repr(value), False))
            elif parsed is not None and (parsed < 0 or parsed > 100):
                audit.append(issue("cover outside 0–100", "Community_level_data", name, record["SampleID"], "error", str(parsed), False))

    presence_sheet = book["inwork_Sp_presence"]
    presence_rows = list(presence_sheet.iter_rows(values_only=True))
    presence_headers = list(presence_rows[0])
    presence_by_sample = {row[4]: dict(zip(presence_headers, row)) for row in presence_rows[1:]}
    for record in records:
        other = presence_by_sample.get(record["SampleID"])
        if not other:
            audit.append(issue("sample absent from presence sheet", "inwork_Sp_presence", None, record["SampleID"], "error", "No cross-check row", False))
            continue
        for name in set(cover_names) & set(presence_headers[8:]):
            cover_present = (number(record.get(name)) or 0) > 0
            presence_present = (number(other.get(name)) or 0) > 0
            if cover_present != presence_present:
                audit.append(issue("presence recorded without numeric cover", "Community_level_data ↔ inwork_Sp_presence", name, record["SampleID"], "info", "Occurrence uses the presence sheet; quantitative cover is unavailable for this plot.", True))

    # Presence is the authoritative occurrence table; cover remains authoritative
    # for abundance. This distinction retains detections that have no numeric cover.
    for name, summary in summaries.items():
        if name not in presence_headers[8:]:
            continue
        habitats: dict[str, Any] = {}
        sites: dict[str, Any] = {}
        for code in HABITATS:
            subset = [row for row in presence_rows[1:] if row[3] == code]
            occupied = sum(1 for row in subset if (number(row[presence_headers.index(name)]) or 0) > 0)
            total = len(subset)
            cover_summary = summary["habitats"].get(code, {}).get("coverAmongOccupiedPlots", summarize_values([]))
            habitats[code] = {
                "sampled": True,
                "occupiedPlots": occupied,
                "totalPlots": total,
                "occurrenceFrequency": round(occupied / total, 6),
                "coverAmongOccupiedPlots": cover_summary,
                "zeroPlotsExcludedFromCoverSummary": True,
                "occurrenceSourceSheet": "inwork_Sp_presence",
                "coverSourceSheet": "Community_level_data",
            }
        for site in SITES:
            subset = [row for row in presence_rows[1:] if row[2] == site]
            occupied = sum(1 for row in subset if (number(row[presence_headers.index(name)]) or 0) > 0)
            sites[site] = {
                "recorded": occupied > 0,
                "occupiedPlots": occupied,
                "totalPlots": len(subset),
                "occurrenceFrequency": round(occupied / len(subset), 6),
            }
        summary["habitats"] = habitats
        summary["sites"] = sites
        summary["occupiedPlots"] = sum(item["occupiedPlots"] for item in habitats.values())
        summary["totalPlots"] = len(presence_rows) - 1

    metadata = {
        "source": workbook_path.name,
        "sheets": {
            "cover": "Community_level_data",
            "presenceCrossCheck": "inwork_Sp_presence",
            "CCI": "CCI_medians_perSpecies",
            "LA": "LA_medians_perSpecies",
        },
        "plotCount": len(records),
        "habitatPlotCounts": dict(plot_totals),
        "sitePlotCounts": dict(site_totals),
        "coverTaxonCount": len(cover_names),
        "studyTaxonCount": len(study_names),
        "nonTaxonColumnsExcluded": sorted(NON_TAXA),
        "calculatedAt": RETRIEVED_AT,
        "methodVersion": "1.0.0",
        "coverMethod": "Occurrence uses inwork_Sp_presence. Cover median, IQR, mean and maximum use positive numeric Community_level_data values and exclude zero/absence plots.",
        "traitMethod": "Workbook pool-wise species summaries retained only when n >= 5; values are 2024 study measurements.",
    }
    return study_names, summaries, audit, metadata


def crosscheck_basic_graphs(main_workbook: Path, supporting_workbook: Path) -> dict[str, Any]:
    """Compare the supplied aggregate workbook with the same sheet in the study workbook."""
    main_book = load_workbook(main_workbook, read_only=True, data_only=True)
    support_book = load_workbook(supporting_workbook, read_only=True, data_only=True)
    main_rows = list(main_book["BasicGraphs"].iter_rows(values_only=True))
    support_rows = list(support_book["BasicGraphs"].iter_rows(values_only=True))
    row_count = max(len(main_rows), len(support_rows))
    column_count = max(max((len(row) for row in main_rows), default=0), max((len(row) for row in support_rows), default=0))
    mismatches = []
    for row_index in range(row_count):
        main_row = main_rows[row_index] if row_index < len(main_rows) else ()
        support_row = support_rows[row_index] if row_index < len(support_rows) else ()
        for column_index in range(column_count):
            main_value = main_row[column_index] if column_index < len(main_row) else None
            support_value = support_row[column_index] if column_index < len(support_row) else None
            if main_value != support_value:
                mismatches.append({"cell": f"R{row_index + 1}C{column_index + 1}", "main": main_value, "support": support_value})
    return {
        "source": supporting_workbook.name,
        "sheet": "BasicGraphs",
        "rows": row_count,
        "columns": column_count,
        "mismatchCount": len(mismatches),
        "matchesMainWorkbook": not mismatches,
        "firstMismatches": mismatches[:20],
    }


def issue(name: str, sheet: str, species: str | None, sample: str | None, severity: str, action: str, resolved: bool) -> dict[str, Any]:
    return {"issue": name, "sourceSheet": sheet, "species": species, "sampleId": sample, "severity": severity, "action": action, "resolved": resolved}


def fetch_taxon_bundle(taxon_id: str) -> tuple[str, dict[str, Any]]:
    taxon = fetch_json(f"taxa/{taxon_id}")
    try:
        descriptions = fetch_json(f"taxa/{taxon_id}/descriptions", {"lang": "en"})
    except RuntimeError:
        descriptions = []
    try:
        media = fetch_json(f"taxa/{taxon_id}/media")
    except RuntimeError:
        media = []
    return taxon_id, {"taxon": taxon, "descriptions": descriptions, "media": media, "retrievedAt": RETRIEVED_AT}


def public_media(item: dict[str, Any], taxon_id: str) -> dict[str, Any] | None:
    abbreviation = item.get("licenseAbbreviation")
    owner = item.get("copyrightOwner") or item.get("author")
    url = item.get("largeURL") or item.get("fullURL")
    if abbreviation not in LICENSES or not owner or not url:
        return None
    license_name, license_url = LICENSES[abbreviation]
    return {
        "imageId": item.get("id"),
        "url": url,
        "thumbnailUrl": item.get("thumbnailURL") or item.get("squareThumbnailURL"),
        "sourcePage": f"https://laji.fi/en/taxon/{taxon_id}/identification",
        "creator": item.get("author") or owner,
        "copyrightOwner": owner,
        "license": license_name,
        "licenseUrl": license_url,
        "attribution": f"Photograph by {owner}, {license_name}, via FinBIF",
        "finbifMediaId": item.get("id"),
        "retrievedAt": RETRIEVED_AT,
        "nonCommercial": "-NC-" in abbreviation,
        "shareAlike": "-SA-" in abbreviation,
    }


def slugify(value: str) -> str:
    return "-".join("".join(character.lower() if character.isalnum() else " " for character in value).split())


def classification(taxon: dict[str, Any]) -> list[dict[str, Any]]:
    parents = taxon.get("parent") or {}
    order = ("kingdom", "phylum", "class", "order", "family", "genus", "species")
    return [
        {"rank": rank, "name": parents[rank].get("scientificName"), "taxonId": parents[rank].get("id"), "authorship": parents[rank].get("scientificNameAuthorship")}
        for rank in order if rank in parents and parents[rank].get("scientificName")
    ]


def download_atlas_media(cache: dict[str, Any], reconciliations: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    atlas = json.loads((CONTENT / "species.json").read_text())
    atlas_ids = {record["taxonId"]: record["slug"] for record in atlas}
    for record in reconciliations:
        if record.get("taxonId") and record["matchStatus"] in {"exact", "accepted-name-match", "synonym-match", "spelling-normalization"}:
            atlas_ids.setdefault(record["taxonId"], slugify(record["finbifAcceptedName"]))
    selected: dict[str, list[dict[str, Any]]] = {}
    for taxon_id, slug in atlas_ids.items():
        bundle = cache.get(taxon_id)
        if not bundle:
            continue
        eligible = [media for item in bundle["media"] if (media := public_media(item, taxon_id))]
        preferred = PRIMARY_MEDIA_IDS.get(taxon_id)
        eligible.sort(key=lambda media: media["imageId"] != preferred)
        eligible = eligible[:3]
        local: list[dict[str, Any]] = []
        for media in eligible:
            media_id = media["finbifMediaId"].replace(".", "-").lower()
            relative = f"/species/{slug}/finbif-{media_id}.webp"
            output = ROOT / "public" / relative.lstrip("/")
            output.parent.mkdir(parents=True, exist_ok=True)
            if not output.exists():
                encoded_url = urllib.parse.quote(media["url"], safe=":/?=&%")
                request = urllib.request.Request(encoded_url, headers={"User-Agent": "Remote-Sensing-Scientist-Academy/1.0"})
                with urllib.request.urlopen(request, timeout=60) as response:
                    image = Image.open(BytesIO(response.read())).convert("RGB")
                image.thumbnail((1400, 1400), Image.Resampling.LANCZOS)
                image.save(output, "WEBP", quality=84, method=6)
            media["file"] = relative
            local.append(media)
        selected[taxon_id] = local
    return selected


def write_outputs(study_names: list[str], reconciliations: list[dict[str, Any]], summaries: dict[str, Any], audit: list[dict[str, Any]], metadata: dict[str, Any], cache: dict[str, Any]) -> None:
    CONTENT.mkdir(parents=True, exist_ok=True)
    DATA.mkdir(parents=True, exist_ok=True)
    DOCS.mkdir(parents=True, exist_ok=True)
    (CONTENT / "taxon-reconciliation.json").write_text(json.dumps(reconciliations, indent=2, ensure_ascii=False) + "\n")
    (DATA / "study-species-summary.json").write_text(json.dumps({"metadata": metadata, "species": summaries}, indent=2, ensure_ascii=False) + "\n")
    selected_media = download_atlas_media(cache, reconciliations)
    normalized_cache = {}
    for taxon_id, bundle in cache.items():
        taxon = bundle["taxon"]
        normalized_cache[taxon_id] = {
            "taxonId": taxon_id,
            "scientificName": taxon.get("scientificName"),
            "scientificNameAuthorship": taxon.get("scientificNameAuthorship"),
            "taxonRank": (taxon.get("taxonRank") or "").removeprefix("MX."),
            "classification": classification(taxon),
            "commonNames": next((r.get("commonNames") for r in reconciliations if r.get("taxonId") == taxon_id and r.get("commonNames")), {}),
            "finland": {
                "occurrenceInFinland": taxon.get("occurrenceInFinland"),
                "latestRedListStatusFinland": taxon.get("latestRedListStatusFinland"),
            },
            "descriptionSections": bundle["descriptions"],
            "media": selected_media.get(taxon_id) or [m for item in bundle["media"] if (m := public_media(item, taxon_id))][:4],
            "primaryImageId": PRIMARY_MEDIA_IDS.get(taxon_id),
            "retrievedAt": bundle["retrievedAt"],
            "sourceUrl": f"https://laji.fi/en/taxon/{taxon_id}",
            "rights": taxon.get("intellectualRights"),
        }
    (DATA / "finbif-cache.json").write_text(json.dumps(normalized_cache, indent=2, ensure_ascii=False) + "\n")
    (DATA / "study-data-audit.json").write_text(json.dumps(audit, indent=2, ensure_ascii=False) + "\n")
    write_reconciliation_report(reconciliations)
    write_study_audit(audit, metadata, reconciliations)
    write_finbif_report(normalized_cache)
    write_image_audit(cache)


def write_reconciliation_report(records: list[dict[str, Any]]) -> None:
    counts = Counter(r["matchStatus"] for r in records)
    lines = ["# Taxon reconciliation report", "", f"Generated {RETRIEVED_AT}. Original workbook labels are preserved. FinBIF fuzzy matches were never accepted automatically.", "", "## Summary", ""]
    lines.extend(f"- {key}: {value}" for key, value in sorted(counts.items()))
    lines += ["", "| Study name | Accepted FinBIF name | Taxon ID | Status | Notes |", "|---|---|---|---|---|"]
    for r in records:
        lines.append(f"| `{r['studyName']}` | {r.get('finbifAcceptedName') or '—'} | {r.get('taxonId') or '—'} | {r['matchStatus']} | {r.get('matchNotes') or '—'} |")
    (DOCS / "taxon-reconciliation-report.md").write_text("\n".join(lines) + "\n")


def write_study_audit(audit: list[dict[str, Any]], metadata: dict[str, Any], reconciliations: list[dict[str, Any]]) -> None:
    unresolved = [r for r in reconciliations if r["matchStatus"] in {"manual-review", "unresolved"}]
    rows = audit + [issue("taxonomy requires review", "inwork_TOTAL_SPECIES", r["studyName"], None, "warning", r["matchNotes"], False) for r in unresolved]
    lines = ["# Study data audit", "", f"Generated {RETRIEVED_AT} from `{metadata['source']}`.", "", f"Plots: {metadata['plotCount']}. Habitat denominators: " + ", ".join(f"{k} {v}" for k, v in metadata["habitatPlotCounts"].items()) + ".", "", "| Issue | Source sheet | Species | Sample ID | Severity | Action | Resolved |", "|---|---|---|---|---|---|---|"]
    if not rows:
        lines.append("| No issues detected | — | — | — | — | — | yes |")
    for r in rows:
        lines.append(f"| {r['issue']} | {r['sourceSheet']} | {r.get('species') or '—'} | {r.get('sampleId') or '—'} | {r['severity']} | {r['action']} | {'yes' if r['resolved'] else 'no'} |")
    (DOCS / "study-data-audit.md").write_text("\n".join(lines) + "\n")


def write_finbif_report(cache: dict[str, Any]) -> None:
    lines = ["# FinBIF import report", "", f"Refreshed {RETRIEVED_AT} from the official FinBIF/Laji.fi public service. FINBIF_ACCESS_TOKEN was {'available' if os.getenv('FINBIF_ACCESS_TOKEN') else 'not available; the public Laji.fi proxy was used'}.", "", "Imported fields: accepted scientific name, authorship, rank, hierarchy, Finnish/Swedish common names when returned, Finland-specific status fields, description-source payloads, and per-image rights metadata.", "", "Description text was cached for provenance but is not machine-translated or automatically republished.", "", f"Resolved taxon records cached: {len(cache)}."]
    (DOCS / "finbif-import-report.md").write_text("\n".join(lines) + "\n")


def write_image_audit(cache: dict[str, Any]) -> None:
    lines = ["# FinBIF image licence audit", "", f"Generated {RETRIEVED_AT}. Publication eligibility requires a reusable mapped licence, creator/owner, URL and source taxon page.", "", "| Taxon ID | Media ID | Creator or owner | Licence | Decision | Reason |", "|---|---|---|---|---|---|"]
    for taxon_id, bundle in sorted(cache.items()):
        for item in bundle["media"]:
            accepted = public_media(item, taxon_id)
            owner = item.get("copyrightOwner") or item.get("author") or "—"
            license_value = item.get("licenseAbbreviation") or "—"
            reason = "Complete reusable rights metadata" if accepted else "Missing creator/URL or licence is absent, unmapped, or not reusable"
            lines.append(f"| {taxon_id} | {item.get('id') or '—'} | {owner} | {license_value} | {'eligible' if accepted else 'excluded'} | {reason} |")
    (DOCS / "image-license-audit.md").write_text("\n".join(lines) + "\n")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--basic-graphs", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--cache-only", action="store_true")
    args = parser.parse_args()
    study_names, summaries, audit, metadata = load_study(args.workbook)
    if args.basic_graphs:
        metadata["basicGraphsCrossCheck"] = crosscheck_basic_graphs(args.workbook, args.basic_graphs)
        if metadata["basicGraphsCrossCheck"]["mismatchCount"]:
            audit.append(issue("BasicGraphs aggregate mismatch", "BasicGraphs", None, None, "warning", "Review the supporting aggregate workbook before using its summaries.", False))
    else:
        metadata["basicGraphsCrossCheck"] = {"performed": False}
    with ThreadPoolExecutor(max_workers=8) as pool:
        reconciliations = list(pool.map(resolve_name, study_names))
    atlas_ids = {record["taxonId"] for record in json.loads((CONTENT / "species.json").read_text())}
    resolved_ids = sorted({r["taxonId"] for r in reconciliations if r.get("taxonId")} | atlas_ids)
    with ThreadPoolExecutor(max_workers=8) as pool:
        cache = dict(pool.map(fetch_taxon_bundle, resolved_ids))
    result = {
        "studyTaxa": len(study_names),
        "statuses": dict(Counter(r["matchStatus"] for r in reconciliations)),
        "resolvedTaxa": len(resolved_ids),
        "CCI": sum("CCI" in s["traits"] for s in summaries.values()),
        "LA": sum("LA" in s["traits"] for s in summaries.values()),
        "auditIssues": len(audit),
    }
    print(json.dumps(result, indent=2))
    if not args.dry_run:
        write_outputs(study_names, reconciliations, summaries, audit, metadata, cache)


if __name__ == "__main__":
    main()
